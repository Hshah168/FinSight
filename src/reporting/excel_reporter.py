"""
FinSight Excel Report Generator
===============================

Output:
- Executive Summary
- P&L Statement
- Variance Analysis
- KPI Metrics

Location:
outputs/reports/
"""

import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


class ExcelReporter:

    def __init__(self):

        self.output_dir = "outputs/reports"

        os.makedirs(self.output_dir, exist_ok=True)

        self.colors = {
            "header": "1F4E78",
            "positive": "C6EFCE",
            "negative": "FFC7CE"
        }

        print("ExcelReporter initialized")

    # MAIN FUNCTION

    def create_monthly_package(
        self,
        pnl_df,
        variance_df,
        metrics_df,
        filename=None
    ):
        if filename is None:
            filename = f"{self.output_dir}/FinSight_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        print(f"Creating Excel report: {filename}")

        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            self._write_sheet(writer, pnl_df, "P&L Statement")
            self._write_sheet(writer, variance_df, "Variance Analysis")
            self._write_sheet(writer, metrics_df, "Key Metrics")

            # Executive Summary: Calculate YoY totals and growth for each metric
            if metrics_df is not None and not metrics_df.empty:
                # Ensure periods are columns, metrics are rows
                df = metrics_df.copy()
                df.index = df.index.astype(str)
                cols_2023 = [col for col in df.columns if str(col).startswith('2023')]
                cols_2024 = [col for col in df.columns if str(col).startswith('2024')]
                totals_2023 = df[cols_2023].sum(axis=1)
                totals_2024 = df[cols_2024].sum(axis=1)
                yoy_growth = ((totals_2024 - totals_2023) / totals_2023.replace(0, float('nan'))) * 100
                summary = pd.DataFrame({
                    'Metric': df.index,
                    '2023 Total': totals_2023.values,
                    '2024 Total': totals_2024.values,
                    'YoY Growth %': yoy_growth.values
                })
            else:
                summary = pd.DataFrame({
                    'Metric': ['No data'],
                    '2023 Total': [0],
                    '2024 Total': [0],
                    'YoY Growth %': [0]
                })
            summary.to_excel(writer, sheet_name="Executive Summary", index=False)

    # SHEET WRITER
    def _write_sheet(
        self,
        writer,
        df,
        sheet_name
    ):
        if df is None or df.empty:
            df = pd.DataFrame({"Info": ["No data available"]})
        # Special handling for Key Metrics: transpose so metric names are first column
        if sheet_name == "Key Metrics" and not df.empty:
            df = df.transpose().reset_index()
            df.columns = ["Metric"] + [str(col) for col in df.columns[1:]]
        # Special handling for P&L Statement: reset index so account details are columns
        if sheet_name == "P&L Statement" and not df.empty and isinstance(df.index, pd.MultiIndex):
            df = df.reset_index()
        df.to_excel(
            writer,
            sheet_name=sheet_name,
            index=False
        )


    # FORMAT ENTIRE WORKBOOK
    def _format_workbook(self, filename):

        wb = openpyxl.load_workbook(filename)

        for ws in wb.worksheets:

            self._format_sheet(ws)

        wb.save(filename)


    # FORMAT SINGLE SHEET
    def _format_sheet(self, ws):

        header_fill = PatternFill(
            start_color=self.colors["header"],
            end_color=self.colors["header"],
            fill_type="solid"
        )

        header_font = Font(
            bold=True,
            color="FFFFFF"
        )

        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # Format header
        for cell in ws[1]:

            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        # Format data cells
        for row in ws.iter_rows(min_row=2):

            for cell in row:

                cell.border = border

                if isinstance(cell.value, (int, float)):

                    if abs(cell.value) > 1:

                        cell.number_format = "$#,##0.00"

                    else:

                        cell.number_format = "0.00%"

        # Auto width
        for column in ws.columns:

            max_length = 0

            column_letter = get_column_letter(column[0].column)

            for cell in column:

                try:

                    if cell.value:

                        max_length = max(
                            max_length,
                            len(str(cell.value))
                        )

                except:
                    pass

            ws.column_dimensions[column_letter].width = min(max_length + 3, 40)